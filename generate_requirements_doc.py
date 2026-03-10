"""
Script tạo tài liệu mô tả nghiệp vụ sản phẩm (Excel)
Dự án: WINDBOT - AI Chatbot tua-bin gió (Qualcomm VR Lab)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "Tai_lieu_mo_ta_nghiep_vu_san_pham.xlsx"

# ── Màu sắc ──────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F4E79"   # xanh đậm — header
COLOR_HEADER_FG   = "FFFFFF"   # trắng — chữ header
COLOR_GROUP_BG    = "D9E1F2"   # xanh nhạt — nhóm title
COLOR_GROUP_FG    = "1F4E79"   # xanh đậm — chữ nhóm
COLOR_MET         = "C6EFCE"   # xanh lá nhạt — đã đáp ứng
COLOR_PARTIAL     = "FFEB9C"   # vàng nhạt — một phần
COLOR_NOT_MET     = "FFC7CE"   # đỏ nhạt — chưa có
COLOR_ROW_ODD     = "FFFFFF"   # trắng
COLOR_ROW_EVEN    = "F5F5F5"   # xám rất nhạt

STATUS_MET     = "✅ Đã đáp ứng"
STATUS_PARTIAL = "⚠️ Một phần"
STATUS_NOT_MET = "❌ Chưa có"

# ── Dữ liệu 19 yêu cầu ────────────────────────────────────────────────────────
# Mỗi mục: (nhóm, yêu_cầu, trạng_thái, giải_pháp, ghi_chú)
REQUIREMENTS = [
    # ── a) Phân tích yêu cầu ─────────────────────────────────────────────────
    (
        "a) Phân tích yêu cầu",
        "Thu thập và chuẩn hóa tài liệu kỹ thuật tua-bin gió (FAQs, sơ đồ, linh kiện, quy trình vận hành/bảo trì, v.v.)",
        STATUS_PARTIAL,
        (
            "Thực trạng: Hệ thống ingestion tài liệu đã hoàn chỉnh — hỗ trợ định dạng PDF, DOCX, PPTX, XLSX, TXT, MD, CSV thông qua LlamaParse (OCR Tiếng Anh + Tiếng Việt) và SimpleDirectoryReader. Metadata tài liệu được lưu vào bảng documents_metadata trên Supabase.\n\n"
            "Còn thiếu: Chưa có bộ dataset tài liệu thực tế về tua-bin gió (FAQs, sơ đồ kỹ thuật, hướng dẫn vận hành). Đây là nội dung dữ liệu, không phải phần mềm.\n\n"
            "Cần làm: Thu thập tài liệu từ EPU/TAF, chuẩn hóa định dạng, nạp vào hệ thống qua pipeline ingest_docs.py."
        ),
        "Phụ thuộc vào nguồn tài liệu từ đối tác EPU/TAF cung cấp. Hạ tầng kỹ thuật đã sẵn sàng tiếp nhận.",
    ),
    (
        "a) Phân tích yêu cầu",
        "Xác định các kịch bản Q&A chính và phân tích nội dung đào tạo liên quan đến tua-bin gió: cấu trúc, linh kiện, vận hành, bảo trì, an toàn, xử lý sự cố",
        STATUS_PARTIAL,
        (
            "Thực trạng: System prompt chuyên biệt cho domain tua-bin gió đã được xây dựng (backend/app/prompts/system.py), hỗ trợ song ngữ Việt-Anh. RAG engine có khả năng trả lời câu hỏi dựa trên tài liệu được nạp. Tính năng gợi ý 3 câu hỏi tiếp theo (follow-up suggestions) đã hoạt động.\n\n"
            "Còn thiếu: Chưa có Q&A corpus chính thức phân loại theo chủ đề (cấu trúc, vận hành, bảo trì, an toàn, sự cố). Chưa có bộ test Q&A để đánh giá chất lượng.\n\n"
            "Cần làm: Xây dựng bộ Q&A mẫu ~100–200 cặp cho mỗi chủ đề; tích hợp làm benchmark đánh giá RAG."
        ),
        "Nên phối hợp với chuyên gia vận hành tua-bin gió của EPU/TAF để xây dựng Q&A corpus chất lượng.",
    ),
    (
        "a) Phân tích yêu cầu",
        "Đánh giá khả năng tích hợp chatbot với Web và VR360",
        STATUS_PARTIAL,
        (
            "Thực trạng: Tích hợp Web đã hoàn chỉnh — Next.js 16 frontend với streaming SSE, responsive design, dark/light mode. API backend FastAPI đã mở (OpenAPI docs tại /docs). CORS đã cấu hình whitelist.\n\n"
            "Còn thiếu: Chưa có đánh giá kỹ thuật chính thức và POC (Proof of Concept) cho VR360. Chưa có tài liệu phân tích feasibility VR.\n\n"
            "Cần làm: Lập báo cáo đánh giá tích hợp VR360, thử nghiệm WebXR API / A-Frame với headset Meta Quest hoặc Pico."
        ),
        "Web integration: HOÀN CHỈNH. VR360: cần nghiên cứu thêm — đề xuất dùng WebXR API (browser-based, không cần cài app riêng).",
    ),

    # ── b) Thiết kế & Phát triển Chatbot ─────────────────────────────────────
    (
        "b) Thiết kế & Phát triển Chatbot",
        "Xây dựng AI chatbot sử dụng công nghệ NLP để hiểu và phản hồi tự nhiên bằng Tiếng Việt và Tiếng Anh",
        STATUS_MET,
        (
            "Thực trạng: Đã đáp ứng đầy đủ — GPT-4o-mini (OpenAI) làm LLM chính với system prompt song ngữ Việt-Anh chuyên biệt cho domain tua-bin gió. Giao diện có nút chuyển ngôn ngữ (vi/en) với persistence LocalStorage. OCR khi ingestion hỗ trợ cả Tiếng Việt và Tiếng Anh (LlamaParse). Chat history được duy trì trong session với memory buffer 4000 tokens.\n\n"
            "Các tính năng NLP đã có: RAG (Retrieval-Augmented Generation), streaming response qua SSE, follow-up suggestions, 2 chế độ chat engine (context và condense_plus_context)."
        ),
        "Sử dụng GPT-4o-mini qua OpenAI API commercial. Không cần fine-tune model riêng ở giai đoạn này.",
    ),
    (
        "b) Thiết kế & Phát triển Chatbot",
        "Xây dựng kho kiến thức chuyên biệt về tua-bin gió (thuật ngữ, quy trình, sơ đồ kỹ thuật)",
        STATUS_PARTIAL,
        (
            "Thực trạng: Hạ tầng kho kiến thức đã hoàn chỉnh — Supabase pgvector (1536 chiều, text-embedding-3-small), collection 'wind_turbine_docs', LlamaIndex VectorStoreIndex với cosine similarity search. Hỗ trợ render sơ đồ Mermaid tương tác (zoom/pan/drag) và công thức LaTeX/KaTeX trong giao diện.\n\n"
            "Còn thiếu: Chưa nạp dataset thực tế về tua-bin gió (thuật ngữ chuyên ngành, quy trình chuẩn, sơ đồ kỹ thuật). Kho kiến thức hiện tại trống hoặc chỉ có dữ liệu test.\n\n"
            "Cần làm: Thu thập và nạp tài liệu chuyên ngành; xây dựng glossary thuật ngữ tua-bin gió Việt-Anh."
        ),
        "Chunk size: 1024 tokens, overlap: 200. Thay đổi chunk size nếu tài liệu kỹ thuật có cấu trúc đặc biệt (bảng, sơ đồ dày đặc).",
    ),
    (
        "b) Thiết kế & Phát triển Chatbot",
        "Thiết kế giao diện thân thiện, tương thích với web, app và VR",
        STATUS_PARTIAL,
        (
            "Thực trạng: Giao diện web đã hoàn chỉnh và thân thiện — Next.js 16 + Tailwind CSS 4, responsive mobile-first, dark/light mode, Markdown/LaTeX/Mermaid rendering, loading indicators, copy button, auto-scroll. Deploy trên Vercel.\n\n"
            "Còn thiếu: Chưa có native mobile app (iOS/Android). Chưa có VR interface. Chưa có Progressive Web App (PWA) manifest.\n\n"
            "Cần làm: (1) Thêm PWA manifest để cài web app trên mobile. (2) Phát triển VR UI overlay cho Meta Quest/Pico. (3) Tùy chọn: React Native app nếu cần native experience."
        ),
        "Ưu tiên PWA trước (ít tốn công nhất, dùng được trên cả iOS/Android). Native app và VR UI là giai đoạn tiếp theo.",
    ),

    # ── c) Quản lý dữ liệu và bảo mật ────────────────────────────────────────
    (
        "c) Quản lý dữ liệu và bảo mật",
        "Phát triển chiến lược dữ liệu: thu thập, chuẩn hóa, gán nhãn và cập nhật kiến thức tua-bin gió",
        STATUS_PARTIAL,
        (
            "Thực trạng: Pipeline ingestion có metadata labeling tự động (filename, file_type, language, num_chunks, ingested_at). Script ingest_docs.py hỗ trợ batch processing theo thư mục. Hỗ trợ 3 tier ingestion (cost_effective, agentic, agentic_plus).\n\n"
            "Còn thiếu: Chưa có tài liệu chiến lược dữ liệu chính thức. Chưa có quy trình data governance (ai được phép nạp, review, xóa tài liệu). Chưa có labeling cho chủ đề (vận hành, bảo trì, an toàn...).\n\n"
            "Cần làm: Viết Data Strategy Document; bổ sung trường 'topic/category' vào metadata; xây dựng quy trình review tài liệu trước khi nạp."
        ),
        "Nên định nghĩa taxonomy chủ đề (ontology) trước khi nạp tài liệu hàng loạt để đảm bảo nhất quán.",
    ),
    (
        "c) Quản lý dữ liệu và bảo mật",
        "Xây dựng Dataset Card (nguồn, cấu trúc, quy trình xử lý)",
        STATUS_NOT_MET,
        (
            "Thực trạng: Bảng documents_metadata trên Supabase có các trường cơ bản (filename, file_type, language, title, num_chunks, ingested_at) nhưng chưa đủ để tạo thành Dataset Card chuẩn.\n\n"
            "Còn thiếu: Chưa có Dataset Card document theo chuẩn Hugging Face hoặc tương đương. Thiếu các thông tin: nguồn gốc tài liệu, license/bản quyền, phiên bản dataset, phạm vi bao phủ, hạn chế, thống kê tổng quan.\n\n"
            "Cần làm: Tạo file dataset_card.md hoặc bảng dataset_cards trong DB; điền đầy đủ các trường theo chuẩn; cập nhật mỗi lần nạp batch mới."
        ),
        "Tham khảo chuẩn Hugging Face Dataset Card: https://huggingface.co/docs/hub/datasets-cards — có thể dùng làm template.",
    ),
    (
        "c) Quản lý dữ liệu và bảo mật",
        "Giao tiếp bảo mật qua HTTPS/TLS",
        STATUS_PARTIAL,
        (
            "Thực trạng: Frontend deploy trên Vercel — HTTPS/TLS được cấu hình tự động và miễn phí. CORS trong FastAPI đã whitelist URL frontend cụ thể (không dùng wildcard *). Supabase connection sử dụng HTTPS mặc định.\n\n"
            "Còn thiếu: Backend API server (FastAPI/Uvicorn) khi deploy production cần SSL certificate riêng nếu không đứng sau reverse proxy. Chưa có HSTS headers. Chưa có certificate pinning.\n\n"
            "Cần làm: Đặt FastAPI backend sau Nginx/Caddy với SSL termination; thêm HSTS header; document quy trình renew certificate."
        ),
        "Khi deploy production: dùng Nginx + Let's Encrypt (Certbot) hoặc Caddy (tự động HTTPS) làm reverse proxy cho FastAPI backend.",
    ),
    (
        "c) Quản lý dữ liệu và bảo mật",
        "Thiết lập quy trình cập nhật và kiểm soát versioning dữ liệu",
        STATUS_NOT_MET,
        (
            "Thực trạng: Hiện tại không có cơ chế versioning cho knowledge base. Tài liệu được nạp một chiều (append-only). Không có khả năng rollback, so sánh phiên bản, hoặc track thay đổi theo thời gian.\n\n"
            "Còn thiếu: Version tracking cho document collections, rollback support, changelog, audit log của các lần cập nhật.\n\n"
            "Cần làm: Thêm trường version và replaced_by vào documents_metadata; xây dựng API endpoint để update/deprecate tài liệu; tạo bảng ingestion_log ghi lại lịch sử thay đổi."
        ),
        "Có thể dùng DVC (Data Version Control) cho file-level versioning hoặc tự build lightweight versioning trong Supabase.",
    ),
    (
        "c) Quản lý dữ liệu và bảo mật",
        "Đảm bảo quyền sở hữu dữ liệu thuộc về EPU/TAF",
        STATUS_NOT_MET,
        (
            "Thực trạng: Dữ liệu hiện lưu trên Supabase cloud (do nhóm phát triển quản lý). Chưa có cơ chế kỹ thuật hay pháp lý nào đảm bảo quyền sở hữu dữ liệu cho EPU/TAF.\n\n"
            "Còn thiếu: Thỏa thuận pháp lý về data ownership (DPA — Data Processing Agreement). Hạ tầng self-hosted hoặc data residency tại Việt Nam. Cơ chế export/migration dữ liệu cho EPU/TAF.\n\n"
            "Cần làm: Ký DPA với EPU/TAF; xem xét self-host Supabase hoặc dùng hosting Việt Nam; cung cấp API export toàn bộ dữ liệu cho EPU/TAF."
        ),
        "Đây là yêu cầu pháp lý/hợp đồng, không chỉ kỹ thuật. Cần tư vấn pháp lý và thảo luận với EPU/TAF về hosting location.",
    ),
    (
        "c) Quản lý dữ liệu và bảo mật",
        "Tuân thủ bảo mật: mã hóa dữ liệu (HTTPS/TLS), ISO 27001 hoặc tương đương",
        STATUS_NOT_MET,
        (
            "Thực trạng: HTTPS/TLS cơ bản đã có (xem yêu cầu #9). Tuy nhiên chưa đáp ứng tiêu chuẩn toàn diện: chưa có authentication/authorization (không có login), chưa có audit log, chưa có rate limiting, chưa có input validation chống injection.\n\n"
            "Còn thiếu: User authentication (JWT/OAuth2), role-based access control, audit log đầy đủ, penetration testing, security policy document, ISO 27001 gap analysis.\n\n"
            "Cần làm: (1) Thêm auth layer (JWT hoặc Supabase Auth). (2) Rate limiting trên API. (3) Input sanitization. (4) Thuê đơn vị audit bảo mật. (5) Lập roadmap ISO 27001."
        ),
        "ISO 27001 certification là quá trình dài (6–18 tháng). Trước mắt: implement các security controls cơ bản, sau đó có thể hướng đến SOC 2 Type II hoặc ISO 27001.",
    ),

    # ── d) Tích hợp hệ thống ──────────────────────────────────────────────────
    (
        "d) Tích hợp hệ thống",
        "Tích hợp chatbot với Web và VR360 (Meta Quest, Pico)",
        STATUS_PARTIAL,
        (
            "Thực trạng: Tích hợp Web hoàn chỉnh — Next.js frontend giao tiếp với FastAPI backend qua SSE streaming, API proxy đã cấu hình trong next.config.ts. Deploy Vercel + backend server.\n\n"
            "Còn thiếu: VR360 integration chưa có. Chưa có SDK/bridge cho Meta Quest hoặc Pico.\n\n"
            "Cần làm: (1) Đánh giá approach: WebXR API (browser-based, dễ nhất) vs Unity XR SDK (native performance tốt hơn) vs A-Frame (declarative HTML-like). (2) Xây dựng POC VR scene với wind turbine model. (3) Nhúng chatbot UI dưới dạng VR overlay panel."
        ),
        "Đề xuất approach: WebXR API + Three.js cho web-based VR — chạy được trên trình duyệt Meta Quest và Pico mà không cần cài app riêng.",
    ),
    (
        "d) Tích hợp hệ thống",
        "Chatbot có khả năng nhận biết ngữ cảnh trong VR (sinh viên chọn linh kiện → chatbot giải thích)",
        STATUS_NOT_MET,
        (
            "Thực trạng: Chưa có VR scene và chưa có cơ chế context awareness từ VR. Backend chat API có hỗ trợ metadata trong request nhưng chưa được dùng cho VR context.\n\n"
            "Còn thiếu: VR scene với 3D wind turbine model có interactive components. Event system để khi người dùng click/chọn linh kiện → gửi context đến chatbot. Context injection vào chat session.\n\n"
            "Cần làm: Thiết kế event schema (component_id, component_name, action) → gửi qua REST API kèm chat message; backend xử lý context injection vào RAG query; VR frontend gửi event khi chọn linh kiện."
        ),
        "Đây là tính năng differentiator quan trọng nhất của dự án. Cần thiết kế API schema cẩn thận để VR engine và chatbot backend giao tiếp hiệu quả.",
    ),
    (
        "d) Tích hợp hệ thống",
        "Đảm bảo hiệu năng: FPS ≥72Hz, VR latency ≤0.5s",
        STATUS_NOT_MET,
        (
            "Thực trạng: VR chưa được implement nên không thể đo FPS và VR latency. Web chat latency hiện tại: ~1–3s đến first token (phụ thuộc OpenAI API). Streaming SSE giảm perceived latency cho người dùng.\n\n"
            "Còn thiếu: VR rendering engine chưa có. Chưa có benchmark FPS. Chưa có optimization cho VR latency.\n\n"
            "Cần làm: Sau khi có VR POC, đo baseline FPS và latency; optimize: (1) Pre-render VR scene assets. (2) Cache common chatbot responses. (3) Edge deployment cho backend gần user. (4) Websocket thay SSE để giảm latency."
        ),
        "FPS ≥72Hz là yêu cầu của VR rendering engine (Three.js/Unity), độc lập với chatbot. VR latency ≤0.5s cho chatbot response rất thách thức — cần cache và streaming response.",
    ),
    (
        "d) Tích hợp hệ thống",
        "API/SDK mở để thuận tiện mở rộng và nâng cấp",
        STATUS_MET,
        (
            "Thực trạng: FastAPI tự động sinh OpenAPI 3.0 documentation tại /docs (Swagger UI) và /redoc (ReDoc). Các endpoint hiện có: POST /api/chat (streaming), POST /api/chat/sessions, GET /api/chat/sessions/{id}/messages, POST /api/ingest, GET /api/health.\n\n"
            "Đã đáp ứng: API design RESTful, JSON request/response chuẩn, SSE cho streaming, CORS cấu hình đúng.\n\n"
            "Nâng cấp thêm: Thêm API versioning (/api/v1/...) để backward compatibility khi nâng cấp. Thêm API key authentication. Tạo SDK client wrapper cho TypeScript/Python."
        ),
        "API versioning nên implement sớm trước khi có bên thứ 3 tích hợp, tránh breaking changes sau này.",
    ),

    # ── e) Đào tạo và hỗ trợ ─────────────────────────────────────────────────
    (
        "e) Đào tạo và hỗ trợ",
        "Cung cấp tài liệu kỹ thuật (kiến trúc hệ thống, API, cài đặt, triển khai, hướng dẫn sử dụng) và hướng dẫn cho đội quản lý EPU/TAF",
        STATUS_PARTIAL,
        (
            "Thực trạng: Đã có README.md (tổng quan kỹ thuật, setup instructions, tech stack) và deploy-guide.md (hướng dẫn deploy Vercel + Ngrok). Có supabase_schema.sql để setup database.\n\n"
            "Còn thiếu: Tài liệu API chi tiết cho EPU/TAF (không phải dev docs). Hướng dẫn admin (thêm tài liệu, quản lý session). Video tutorial. Hướng dẫn vận hành hệ thống hàng ngày.\n\n"
            "Cần làm: Viết Admin Guide (Tiếng Việt) hướng dẫn EPU/TAF quản lý chatbot; tạo video walkthrough; tổ chức buổi training hands-on."
        ),
        "Tài liệu cho EPU/TAF nên dùng ngôn ngữ phi-kỹ thuật, tập trung vào workflow hàng ngày (thêm tài liệu mới, xem lịch sử chat, xử lý lỗi thường gặp).",
    ),
    (
        "e) Đào tạo và hỗ trợ",
        "Bàn giao AI model, dataset và training pipeline (nếu có)",
        STATUS_PARTIAL,
        (
            "Thực trạng: Dự án sử dụng OpenAI GPT-4o-mini qua commercial API (không có model fine-tuned riêng). Dataset ingestion pipeline đã có (ingest_docs.py với LlamaParse). Vector embeddings lưu trong Supabase pgvector.\n\n"
            "Còn thiếu: Chưa có gói bàn giao chính thức. Chưa export được vector embeddings hiện tại. Chưa có hướng dẫn migrate sang model khác nếu EPU/TAF muốn dùng open-source LLM.\n\n"
            "Cần làm: Tạo gói bàn giao gồm: (1) Source code toàn bộ. (2) Export dataset đã ingestion (documents + embeddings). (3) Hướng dẫn cấu hình API keys. (4) Script backup/restore Supabase. (5) Tùy chọn: migrate sang Ollama/vLLM nếu cần on-premise."
        ),
        "Vì dùng OpenAI API, EPU/TAF cần account OpenAI và API key riêng khi tiếp nhận vận hành. Nên cân nhắc option self-hosted LLM (Ollama + Llama 3) để tránh phụ thuộc vendor.",
    ),
    (
        "e) Đào tạo và hỗ trợ",
        "Hỗ trợ sau triển khai ít nhất 3 tháng, bao gồm sửa lỗi và cập nhật nhỏ",
        STATUS_NOT_MET,
        (
            "Thực trạng: Không có điều khoản hỗ trợ sau triển khai trong codebase (đây là thỏa thuận hợp đồng/dịch vụ, không phải phần mềm).\n\n"
            "Còn thiếu: SLA document (Service Level Agreement). Kênh tiếp nhận lỗi (issue tracker, email, hotline). Quy trình xử lý bug report. Lịch cập nhật định kỳ. Monitoring và alerting hệ thống.\n\n"
            "Cần làm: (1) Ký phụ lục hợp đồng hỗ trợ 3 tháng với SLA cụ thể. (2) Setup GitHub Issues hoặc Jira cho bug tracking. (3) Cài monitoring (Sentry cho frontend, Loguru/Datadog cho backend). (4) Lịch cập nhật hàng tháng."
        ),
        "Nên định nghĩa SLA rõ ràng: thời gian phản hồi (ví dụ: critical bug ≤4h, normal bug ≤2 ngày làm việc). Đề xuất setup Sentry.io miễn phí cho error tracking tự động.",
    ),
]


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def make_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def make_font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tài liệu nghiệp vụ"

    # ── Tiêu đề tài liệu ─────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "TÀI LIỆU MÔ TẢ NGHIỆP VỤ SẢN PHẨM — AI CHATBOT TUA-BIN GIÓ (WINDBOT)"
    title_cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=14, name="Calibri")
    title_cell.fill = make_fill(COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    sub_cell = ws["A2"]
    sub_cell.value = "Dự án: Qualcomm VR Lab  |  Phiên bản: 1.0  |  Ngày: 10/03/2026  |  Trạng thái: ✅ Đã đáp ứng  |  ⚠️ Một phần  |  ❌ Chưa có"
    sub_cell.font = Font(italic=True, color="595959", size=10, name="Calibri")
    sub_cell.fill = make_fill("EBF3FB")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Header cột ───────────────────────────────────────────────────────────
    headers = ["STT", "Nhóm yêu cầu", "Yêu cầu (Requirement)", "Giải pháp / Thực trạng (Solution)", "Ghi chú (Note)"]
    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = make_font(bold=True, color=COLOR_HEADER_FG, size=11)
        cell.fill = make_fill(COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
    ws.row_dimensions[header_row].height = 28

    # ── Dữ liệu ──────────────────────────────────────────────────────────────
    current_group = None
    data_row = header_row + 1
    stt = 1

    for group, requirement, status, solution, note in REQUIREMENTS:
        # In group header row nếu nhóm mới
        if group != current_group:
            current_group = group
            ws.merge_cells(f"A{data_row}:E{data_row}")
            grp_cell = ws.cell(row=data_row, column=1, value=group.upper())
            grp_cell.font = make_font(bold=True, color=COLOR_GROUP_FG, size=11)
            grp_cell.fill = make_fill(COLOR_GROUP_BG)
            grp_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            grp_cell.border = make_border()
            ws.row_dimensions[data_row].height = 22
            data_row += 1

        # Màu nền theo trạng thái
        if status == STATUS_MET:
            row_fill = make_fill(COLOR_MET)
        elif status == STATUS_PARTIAL:
            row_fill = make_fill(COLOR_PARTIAL)
        else:
            row_fill = make_fill(COLOR_NOT_MET)

        row_data = [stt, "", requirement, f"{status}\n\n{solution}", note]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = make_border()
            cell.alignment = Alignment(
                horizontal="left" if col_idx > 1 else "center",
                vertical="top",
                wrap_text=True,
            )
            if col_idx == 1:
                cell.font = make_font(bold=True, size=11)
            elif col_idx == 4:
                # Tô đậm dòng trạng thái trong solution
                cell.font = make_font(size=10)
            else:
                cell.font = make_font(size=10)

        ws.row_dimensions[data_row].height = 130
        stt += 1
        data_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [6, 28, 50, 72, 42]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes & auto-filter ────────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:E{header_row}"

    # ── Legend sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Chú giải")
    ws2["A1"] = "CHÚ GIẢI MÀU SẮC VÀ TRẠNG THÁI"
    ws2["A1"].font = make_font(bold=True, size=13)
    ws2["A1"].fill = make_fill(COLOR_HEADER_BG)
    ws2["A1"].font = Font(bold=True, color=COLOR_HEADER_FG, size=13, name="Calibri")
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    legend_data = [
        ("Màu nền", "Trạng thái", "Ý nghĩa"),
        ("Xanh lá nhạt", STATUS_MET, "Yêu cầu đã được implement đầy đủ trong dự án hiện tại"),
        ("Vàng nhạt", STATUS_PARTIAL, "Yêu cầu đã implement một phần, còn thiếu hoặc cần bổ sung"),
        ("Đỏ nhạt", STATUS_NOT_MET, "Yêu cầu chưa được implement, cần phát triển từ đầu"),
    ]
    fill_map = {
        legend_data[1][1]: make_fill(COLOR_MET),
        legend_data[2][1]: make_fill(COLOR_PARTIAL),
        legend_data[3][1]: make_fill(COLOR_NOT_MET),
    }

    for r_idx, (col1, col2, col3) in enumerate(legend_data, start=2):
        ws2.cell(row=r_idx, column=1, value=col1).border = make_border()
        ws2.cell(row=r_idx, column=2, value=col2).border = make_border()
        ws2.cell(row=r_idx, column=3, value=col3).border = make_border()
        if r_idx > 2:
            status_key = legend_data[r_idx - 2][1]
            for c in range(1, 4):
                ws2.cell(row=r_idx, column=c).fill = fill_map[status_key]
                ws2.cell(row=r_idx, column=c).font = make_font(size=10)
                ws2.cell(row=r_idx, column=c).alignment = Alignment(wrap_text=True, vertical="center")
        else:
            for c in range(1, 4):
                ws2.cell(row=r_idx, column=c).fill = make_fill("D9E1F2")
                ws2.cell(row=r_idx, column=c).font = make_font(bold=True, size=10)
                ws2.cell(row=r_idx, column=c).alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r_idx].height = 40

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 60

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Tổng kết")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "TỔNG KẾT MỨC ĐỘ ĐÁP ỨNG YÊU CẦU"
    ws3["A1"].font = Font(bold=True, color=COLOR_HEADER_FG, size=13, name="Calibri")
    ws3["A1"].fill = make_fill(COLOR_HEADER_BG)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    count_met     = sum(1 for _, _, s, _, _ in REQUIREMENTS if s == STATUS_MET)
    count_partial = sum(1 for _, _, s, _, _ in REQUIREMENTS if s == STATUS_PARTIAL)
    count_not_met = sum(1 for _, _, s, _, _ in REQUIREMENTS if s == STATUS_NOT_MET)
    total = len(REQUIREMENTS)

    summary_headers = ["Trạng thái", "Số yêu cầu", "Tỷ lệ", "Nhận xét"]
    for c_idx, h in enumerate(summary_headers, 1):
        cell = ws3.cell(row=2, column=c_idx, value=h)
        cell.font = make_font(bold=True, size=10)
        cell.fill = make_fill("D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = make_border()
    ws3.row_dimensions[2].height = 22

    summary_rows = [
        (STATUS_MET,     count_met,     f"{count_met/total*100:.0f}%",     "Sẵn sàng cho production",          COLOR_MET),
        (STATUS_PARTIAL, count_partial, f"{count_partial/total*100:.0f}%", "Cần bổ sung để đáp ứng đầy đủ",    COLOR_PARTIAL),
        (STATUS_NOT_MET, count_not_met, f"{count_not_met/total*100:.0f}%", "Cần phát triển mới hoàn toàn",     COLOR_NOT_MET),
        ("TỔNG CỘNG",   total,          "100%",                            "19 yêu cầu kỹ thuật từ khách hàng", "D9D9D9"),
    ]

    for r_offset, (status, count, pct, remark, bg) in enumerate(summary_rows, start=3):
        for c_idx, val in enumerate([status, count, pct, remark], 1):
            cell = ws3.cell(row=r_offset, column=c_idx, value=val)
            cell.fill = make_fill(bg)
            cell.border = make_border()
            cell.font = make_font(bold=(r_offset == 6), size=10)
            cell.alignment = Alignment(horizontal="center" if c_idx in (1, 2, 3) else "left",
                                       vertical="center", wrap_text=True)
        ws3.row_dimensions[r_offset].height = 28

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 40

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"✅ Đã tạo file: {OUTPUT_FILE}")
    print(f"   Sheet 1: Tài liệu nghiệp vụ ({total} yêu cầu)")
    print(f"   Sheet 2: Chú giải màu sắc")
    print(f"   Sheet 3: Tổng kết — ✅{count_met} | ⚠️{count_partial} | ❌{count_not_met}")


if __name__ == "__main__":
    build_excel()
